import os
import shutil
import argparse
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook

# python.exe .\generate_fdb.py .\fdb.xlsx

# 设置输出路径
output_dir = 'out'

# 复制文件到指定目录（假设 src 和 inc 文件夹在当前目录下）
src_dir = '../src'  # 假设 src 文件夹在当前目录
inc_dir = '../inc'  # 假设 inc 文件夹在当前目录

# 设置 Jinja2 环境，模板文件存放在 templates 文件夹下
env = Environment(
    loader=FileSystemLoader('templates'),
    trim_blocks=True,  # 去除行尾多余的空白
    lstrip_blocks=True  # 去除行首多余的空白
)

# 创建命令行参数解析器
parser = argparse.ArgumentParser(description="根据 Excel 文件生成 .h 和 .c 文件")
parser.add_argument('excel_file', help="输入的 Excel 文件路径，例如 data.xlsx")
args = parser.parse_args()

# 获取 Excel 文件路径
excel_file_path = args.excel_file

# 检查文件是否存在
if not os.path.exists(excel_file_path):
    print(f"错误: 文件 '{excel_file_path}' 不存在！")
    exit(1)

# 读取 Excel 文件
wb = load_workbook(excel_file_path)
ws = wb.active

variables = []

# 读取表格数据，假设数据从第二行开始
for row in ws.iter_rows(min_row=2, values_only=True):
    try:
        # 确保每行数据有4列
        if len(row) != 4:
            print(f"警告: 跳过无效行 {row}")
            continue
        
        var_name, var_type, array_size, macro_name = row
        variables.append({
            'var_name': var_name,
            'type': var_type,
            'array_size': array_size,
            'macro_name': macro_name
        })
    except Exception as e:
        print(f"错误: 无法解析行 {row}. 错误信息: {e}")

# 渲染模板并生成 .h 文件
header_template = env.get_template('flashdb_table.h.j2')
header_content = header_template.render(variables=variables)

# 渲染模板并生成 .c 文件
source_template = env.get_template('flashdb_table.c.j2')
source_content = source_template.render(variables=variables)

# 确保 build 文件夹存在
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# 保存文件到 build 目录
header_file_path = os.path.join(output_dir, 'flashdb_table.h')
source_file_path = os.path.join(output_dir, 'flashdb_table.c')

with open(header_file_path, 'w') as f:
    f.write(header_content)

with open(source_file_path, 'w') as f:
    f.write(source_content)

# 确保目标文件夹存在
if not os.path.exists(src_dir):
    os.makedirs(src_dir)

if not os.path.exists(inc_dir):
    os.makedirs(inc_dir)

# 复制文件到 src 和 inc 文件夹
shutil.copy(header_file_path, os.path.join(inc_dir, 'flashdb_table.h'))
shutil.copy(source_file_path, os.path.join(src_dir, 'flashdb_table.c'))
