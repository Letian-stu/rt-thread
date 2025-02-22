import shutil,os
import datetime

from enum import Enum
from openpyxl import load_workbook
from jinja2 import Environment,FileSystemLoader

class attr(Enum):
    attr_name_cn = 2
    attr_name_en = 3
    attr_type = 4
    attr_mean = 6
    attr_min = 7
    attr_max = 8
    attr_unit = 9
    attr_pid = 10
    attr_addr = 11

class datatype(Enum):
    type_uint8 = 0
    type_int8 = 1
    type_uint16 = 2
    type_int16 = 3
    type_uint32 = 4
    type_int32 = 5
    type_float = 6
    type_mem = 7
    type_max = 8

home_addr = {
    "ATTR_AUTHOR":"B4",
    "ATTR_VERSION":"B3",
    "ATTR_TIME":"B2"
}

attr_addr = {
    "ATTR_UINT8_DATA_NUM":"N2",
    "ATTR_INT8_DATA_NUM":"N3",
    "ATTR_UINT16_DATA_NUM":"N4",
    "ATTR_INT16_DATA_NUM":"N5",
    "ATTR_UINT32_DATA_NUM":"N6",
    "ATTR_INT32_DATA_NUM":"N7",
    "ATTR_FLOAT_DATA_NUM":"N8",
    "ATTR_MEM_DATA_NUM":"N9",
    "ATTR_SUM_DATA_NUM":"N10",
    "ATTR_SUM_DATA_SIZE":"O10"
}

config_addr = {
    "ATTR_UINT8_START_ADDR":"F2",
    "ATTR_INT8_START_ADDR":"F3",
    "ATTR_UINT16_START_ADDR":"F4",
    "ATTR_INT16_START_ADDR":"F5",
    "ATTR_UINT32_START_ADDR":"F6",
    "ATTR_INT32_START_ADDR":"F7",
    "ATTR_FLOAT_START_ADDR":"F8",
    "ATTR_MEM_START_ADDR":"F9",
    "ATTR_MAX_SIZE":"J14",
    "ATTR_MEM_LEN":"C9"
}

g_file_name = "./attr.xlsx"
g_xlsx_table_home = "home"
g_xlsx_table_attr = "attr"
g_xlsx_table_config = "config"

g_attr_defconfig_h = "attr_defconfig.h.j2"
g_attr_pid_table_h = "attr_pid_table.h.j2"

output_path = "./out"
code_path = "../inc"

def openfile(file, table):
    wb = load_workbook(filename = file, data_only=True) #打开文件&只读取表内值
    sheet_ranges = wb[table]
    return sheet_ranges

#拷贝备用文件
def copy_bak_files(filename):   
    if not os.path.exists("./out"):
        os.makedirs("./out")
    else:
        shutil.rmtree("./out") 
        os.makedirs("./out")

    current_time = datetime.datetime.now()
    time_str = current_time.strftime("%m-%d-%H-%M")

    bak_filename = "attr-bak-"+ time_str +".xlsx"   
    shutil.copy(filename , "./out/"+bak_filename)
    print(filename,"copied to",bak_filename)       

#获取属性
def get_attr(attr_val):
    sheet_ranges = openfile(g_file_name, g_xlsx_table_attr)
    for i in range(2,sheet_ranges.max_row + 1):
        attr_dict = {}
        for j in attr:
            attr_dict[j.name] = sheet_ranges.cell(row=i, column=j.value).value
        # print(attr_dict)
        attr_val.append(attr_dict)

#获取宏定义
def get_define(config_val):
    read_list = {
        g_xlsx_table_home:home_addr,
        g_xlsx_table_attr:attr_addr,
        g_xlsx_table_config:config_addr,
    }

    for name, grade in read_list.items():
        sheet_ranges = openfile(g_file_name, name)
        for name, grade in grade.items():
            config_dict = {}
            config_dict[name] = sheet_ranges[grade].value
            config_val.update(config_dict)

#写入文件
def j2_write_tempfile(deffile, var):
    current_time = datetime.datetime.now()
    var["ATTR_TIME"] = current_time.strftime("%Y/%m/%d %H:%M")
    env = Environment(loader = FileSystemLoader('./'))
    tpl = env.get_template(deffile) 
    outfile, suffix = os.path.splitext(deffile) #读取文件名不带后缀
    with open('./out/'+outfile,'w+', encoding='UTF-8') as fout:
        render_content = tpl.render(var)
        fout.write(render_content)      


if __name__ == '__main__':  
        copy_bak_files(g_file_name)         

        config_dict = {}
        get_define(config_dict)
        j2_write_tempfile(g_attr_defconfig_h, config_dict)

        attr_list = [] 
        get_attr(attr_list)
        vars = {
            'attr_list':attr_list
        }
        j2_write_tempfile(g_attr_pid_table_h, vars)





























