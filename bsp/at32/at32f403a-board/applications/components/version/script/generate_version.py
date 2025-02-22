import os
import shutil
from datetime import datetime
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

# ==================== 硬编码配置 ====================
EXCEL_FILE = 'version.xlsx'          # 输入Excel文件路径
OUTPUT_DIR = 'out'                   # 生成的中间文件输出目录
SRC_DIR = '../src'                   # 最终源码文件存放目录
INC_DIR = '../inc'                   # 最终头文件存放目录
ENABLE_COPY = True                   # 是否复制到目标目录
# ==================================================

def generate_version_files():
    """生成版本信息文件"""
    # 创建输出目录
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 读取Excel数据
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['version']
    
    history = []
    required_columns = [0, 1, 2, 3, 4]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not validate_row(row, required_columns):
            continue
            
        try:
            entry = process_row(row)
            history.append(entry)
        except Exception as e:
            print(f"第{row_idx}行处理失败:{str(e)}")
            continue

    # 渲染模板
    env = Environment(
        loader=FileSystemLoader('templates'),
        trim_blocks=True,
        lstrip_blocks=True
    )
    
    # 生成头文件
    header_path = os.path.join(OUTPUT_DIR, 'dev_version.h')
    with open(header_path, 'w') as f:
        f.write(env.get_template('dev_version.h.j2').render(history=history))

    # 生成源文件
    source_path = os.path.join(OUTPUT_DIR, 'dev_version.c')
    with open(source_path, 'w') as f:
        f.write(env.get_template('dev_version.c.j2').render(history=history))

    # 复制到目标目录
    if ENABLE_COPY:
        os.makedirs(SRC_DIR, exist_ok=True)
        os.makedirs(INC_DIR, exist_ok=True)
        
        shutil.copy(header_path, os.path.join(INC_DIR, 'dev_version.h'))
        shutil.copy(source_path, os.path.join(SRC_DIR, 'dev_version.c'))
        print(f"文件已复制到:{INC_DIR}/dev_version.h 和 {SRC_DIR}/dev_version.c")

def validate_row(row, required_cols):
    """验证行数据有效性"""
    for col in required_cols:
        if row[col] in (None, "", " "):
            return False
    return True

def process_row(row):
    """处理单行数据（带容错机制）"""
    # 日期处理（支持多种格式）
    raw_date = row[0]
    try:
        if isinstance(raw_date, datetime):
            dt = raw_date
        else:
            # 尝试解析字符串日期
            dt = parse_flexible_date(raw_date)
        
        date_str = dt.strftime("%Y-%m-%d %H:%M")
    except Exception as e:
        raise ValueError(f"日期解析失败:{raw_date}") from e
    
    # 版本号处理
    version = str(row[2]).strip().lstrip('V')
    if not version.replace('.', '').isdigit():
        raise ValueError(f"非法版本号:{row[2]}")
    
    return {
        'date': date_str,
        'author': row[1].strip(),
        'version': version,
        'change_type': row[3].strip(),
        'description': str(row[4]).strip()
    }

def parse_flexible_date(date_str):
    """智能日期解析"""
    formats = [
        "%d/%m/%y-%H:%M",    # 原始格式
        "%d/%m/%Y-%H:%M",    # 支持4位年份
        "%Y-%m-%d %H:%M",    # ISO格式
        "%m/%d/%y %H:%M"     # 美国格式
    ]
    
    # 修正非法时间（如24:00→00:00）
    if '-24:' in date_str:
        date_str = date_str.replace('-24:', '-00:')
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    raise ValueError(f"无法识别的日期格式:{date_str}")
if __name__ == "__main__":
    # 输入文件校验
    if not os.path.exists(EXCEL_FILE):
        print(f"错误:输入文件 {EXCEL_FILE} 不存在")
        exit(1)

    try:
        generate_version_files()
        print(f"生成完成!输出目录:{os.path.abspath(OUTPUT_DIR)}")
    except Exception as e:
        print(f"生成失败:{str(e)}")
        exit(1)