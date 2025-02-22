from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

# 自定义过滤器：rjust
def rjust_filter(value, width):
    return str(value).rjust(width)

def read_partition_table(file_path):
    """读取Excel中的分区表并打印出来"""
    wb = load_workbook(file_path, data_only=True)  # Open the workbook and get only the evaluated values
    ws = wb.active  # Assuming the relevant data is in the first sheet

    # Store partition data
    onchip_partitions = []
    nor_partitions = []

    # Iterate over the rows and extract partition info
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
        try:
            size_kb = float(row[4]) if row[4] is not None else 0
            start_addr_decimal = row[2]
            start_addr_hex = row[3]

            partition = {
                'dev': row[0],  # Device name ("OnChip" or "w25q64")
                'name': row[1],  # Partition name
                'start_addr': start_addr_decimal,  # Start address
                'start_hex': start_addr_hex,      # Start address in hex
                'size_kb': size_kb,               # Size in KB
                'size_bytes': int(size_kb*1024)       # Size in bytes
            }

            if row[0] == "OnChip":
                onchip_partitions.append(partition)
            elif row[0] == "w25q64":
                nor_partitions.append(partition)

        except (ValueError, TypeError) as e:
            print(f"Skipping invalid row: {row}. Error: {e}")

    return onchip_partitions, nor_partitions

def generate_cfg_file(onchip_partitions, nor_partitions, template_path, output_path):
    """Render the template and generate the fal_cfg.h file"""
    # Set up the Jinja2 environment with whitespace trimming
    env = Environment(
        loader=FileSystemLoader(template_path),
        trim_blocks=True,  # Trim newlines after blocks
        lstrip_blocks=True  # Remove leading whitespaces from blocks
    )
    env.filters['rjust'] = rjust_filter  # 注册自定义过滤器

    template = env.get_template('fal_cfg_template.h.j2')

    # Get the device names from the first partition entry (assuming there is at least one partition)
    onchip_device = onchip_partitions[0]['dev'] if onchip_partitions else 'onchip_flash'
    nor_device = nor_partitions[0]['dev'] if nor_partitions else 'w25q64'

    # Prepare data for template rendering
    data = {
        'onchip_device_macro': onchip_device,
        'nor_device_macro': nor_device,
        'onchip_partitions': onchip_partitions,
        'nor_partitions': nor_partitions
    }

    # Render the template with data and write to the output file
    with open(output_path, 'w') as f:
        f.write(template.render(data))

    print(f"Generated configuration file at {output_path}")

# Usage
file_path = 'fal.xlsx'  # Path to the uploaded Excel file
template_path = 'templates'     # Directory where the template is stored
output_path = 'out/fal_cfg.h'  # Output file path

onchip_partitions, nor_partitions = read_partition_table(file_path)
generate_cfg_file(onchip_partitions, nor_partitions, template_path, output_path)
